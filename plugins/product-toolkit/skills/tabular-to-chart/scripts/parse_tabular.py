#!/usr/bin/env python3
"""
parse_tabular.py
범용 테이블 데이터 파서. CSV와 Excel(.xlsx)을 지원한다.

xlsx 파싱 전략:
  1순위: stdlib (zipfile + xml.etree) — 단순 2D 시트용, 외부 패키지 불필요
  2순위: openpyxl — 병합셀·수식·다중시트 등 복잡한 구조용 (pip install 필요)
  3순위: 실패 시 에러 메시지와 함께 CSV 변환 안내

Usage:
  python3 parse_tabular.py --file data.csv --summary
  python3 parse_tabular.py --file report.xlsx --summary
  python3 parse_tabular.py --file report.xlsx --sheet "Sheet1" --to-csv output.csv
  python3 parse_tabular.py --file report.xlsx --list-sheets
"""

import argparse
import csv
import json
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET


# ── stdlib xlsx 파서 ─────────────────────────────────────────

_NS = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}


def _col_index(col_str: str) -> int:
    """'A' → 0, 'B' → 1, 'AA' → 26 ..."""
    result = 0
    for ch in col_str:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1


def _parse_cell_ref(ref: str) -> tuple[int, int]:
    """'B3' → (row=2, col=1)"""
    m = re.match(r"([A-Z]+)(\d+)", ref)
    if not m:
        return (0, 0)
    return int(m.group(2)) - 1, _col_index(m.group(1))


def _xlsx_list_sheets_stdlib(path: Path) -> list[str]:
    """xlsx 내 시트 이름 목록 (stdlib)"""
    with zipfile.ZipFile(path) as zf:
        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        sheets = wb_xml.findall(".//s:sheet", _NS)
        return [s.get("name", "") for s in sheets]


def _xlsx_to_rows_stdlib(path: Path, sheet_name: str | None = None) -> list[list[str]]:
    """xlsx를 2D 문자열 리스트로 파싱 (stdlib only).
    병합셀·수식 결과값·인라인 문자열은 처리하지 않는다."""
    with zipfile.ZipFile(path) as zf:
        # shared strings
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            ss_xml = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in ss_xml.findall(".//s:si", _NS):
                texts = si.findall(".//s:t", _NS)
                shared.append("".join(t.text or "" for t in texts))

        # 시트 선택
        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        sheets = wb_xml.findall(".//s:sheet", _NS)
        sheet_names = [s.get("name", "") for s in sheets]

        if sheet_name:
            if sheet_name not in sheet_names:
                raise ValueError(f"시트 '{sheet_name}'을 찾을 수 없습니다. 사용 가능: {sheet_names}")
            idx = sheet_names.index(sheet_name)
        else:
            idx = 0

        sheet_path = f"xl/worksheets/sheet{idx + 1}.xml"
        if sheet_path not in zf.namelist():
            # rId 기반 매핑 시도
            rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
            ns_r = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            r_id = sheets[idx].get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
            for rel in rels_xml.findall(".//r:Relationship", ns_r):
                if rel.get("Id") == r_id:
                    sheet_path = "xl/" + rel.get("Target", "")
                    break

        sheet_xml = ET.fromstring(zf.read(sheet_path))

        # 셀 파싱
        max_row = 0
        max_col = 0
        cells: dict[tuple[int, int], str] = {}

        for row_el in sheet_xml.findall(".//s:sheetData/s:row", _NS):
            for cell in row_el.findall("s:c", _NS):
                ref = cell.get("r", "")
                if not ref:
                    continue
                r, c = _parse_cell_ref(ref)
                max_row = max(max_row, r)
                max_col = max(max_col, c)

                val_el = cell.find("s:v", _NS)
                raw = (val_el.text or "") if val_el is not None else ""

                cell_type = cell.get("t", "")
                if cell_type == "s" and raw:
                    try:
                        cells[(r, c)] = shared[int(raw)]
                    except (IndexError, ValueError):
                        cells[(r, c)] = raw
                elif cell_type == "inlineStr":
                    is_el = cell.find(".//s:is", _NS)
                    if is_el is not None:
                        texts = is_el.findall(".//s:t", _NS)
                        cells[(r, c)] = "".join(t.text or "" for t in texts)
                    else:
                        cells[(r, c)] = raw or ""
                else:
                    cells[(r, c)] = raw or ""

        # 2D 리스트 구성
        rows = []
        for r in range(max_row + 1):
            row = []
            for c in range(max_col + 1):
                row.append(cells.get((r, c), ""))
            rows.append(row)

        return rows


# ── openpyxl 폴백 ───────────────────────────────────────────

def _xlsx_list_sheets_openpyxl(path: Path) -> list[str]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _xlsx_to_rows_openpyxl(path: Path, sheet_name: str | None = None) -> list[list[str]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    wb.close()
    return rows


# ── 통합 인터페이스 ─────────────────────────────────────────

def list_sheets(path: Path) -> list[str]:
    if path.suffix.lower() != ".xlsx":
        return []
    try:
        return _xlsx_list_sheets_stdlib(path)
    except Exception:
        try:
            return _xlsx_list_sheets_openpyxl(path)
        except ImportError:
            print("오류: xlsx 시트 목록을 읽을 수 없습니다. openpyxl 설치를 권장합니다.", file=sys.stderr)
            sys.exit(1)


def load_tabular(path: Path, sheet_name: str | None = None) -> list[list[str]]:
    """CSV 또는 xlsx를 2D 문자열 리스트로 로드한다."""
    ext = path.suffix.lower()

    if ext == ".csv":
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            return [row for row in reader]

    if ext in (".xlsx", ".xls"):
        if ext == ".xls":
            print("오류: .xls(구 형식)은 지원하지 않습니다. .xlsx로 다시 저장해 주세요.", file=sys.stderr)
            sys.exit(1)

        # 1순위: stdlib
        stdlib_result = None
        try:
            stdlib_result = _xlsx_to_rows_stdlib(path, sheet_name)
        except Exception as e:
            print(f"  [정보] stdlib 파싱 실패 ({e}), openpyxl 시도...", file=sys.stderr)

        # stdlib 결과가 있어도 헤더 행이 전부 빈 문자열이면 inlineStr 등 미지원 형식일 수 있음
        # → openpyxl 로 재시도한다
        if stdlib_result is not None:
            header_empty = stdlib_result and all(v == "" for v in stdlib_result[0])
            if not header_empty:
                return stdlib_result
            print("  [정보] stdlib 파서가 빈 헤더를 반환했습니다. openpyxl 재시도...", file=sys.stderr)

        # 2순위: openpyxl
        try:
            return _xlsx_to_rows_openpyxl(path, sheet_name)
        except ImportError:
            print("오류: 이 xlsx 파일은 복잡한 구조로, openpyxl이 필요합니다.", file=sys.stderr)
            print("  설치: pip install openpyxl  (venv 권장)", file=sys.stderr)
            sys.exit(1)
        except Exception as e:
            print(f"오류: xlsx 파싱 실패 — {e}", file=sys.stderr)
            sys.exit(1)

    print(f"오류: 지원하지 않는 파일 형식 — {ext}", file=sys.stderr)
    sys.exit(1)


def rows_to_dicts(rows: list[list[str]]) -> list[dict[str, str]]:
    """첫 행을 헤더로 사용하여 딕셔너리 리스트로 변환."""
    if not rows:
        return []
    headers = [h.strip() for h in rows[0]]
    return [{headers[i]: (row[i].strip() if i < len(row) else "") for i in range(len(headers))} for row in rows[1:]]


# ── 구조 분석 ───────────────────────────────────────────────

def _infer_type(values: list[str]) -> str:
    """샘플 값들로 컬럼 타입 추론."""
    nums, pcts, dates, empties = 0, 0, 0, 0
    for v in values:
        v = v.strip()
        if not v:
            empties += 1
            continue
        if re.match(r"^-?\d+\.?\d*$", v.replace(",", "")):
            nums += 1
        elif "%" in v:
            pcts += 1
        elif re.match(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", v):
            dates += 1

    total = len(values) - empties
    if total == 0:
        return "empty"
    if nums / total > 0.7:
        return "numeric"
    if pcts / total > 0.7:
        return "percentage"
    if dates / total > 0.7:
        return "date"
    return "text"


def summarize(rows: list[list[str]]) -> dict:
    """테이블 구조 요약 JSON 생성."""
    if not rows:
        return {"error": "빈 데이터"}

    headers = [h.strip() for h in rows[0]]
    data_rows = rows[1:]
    n_rows = len(data_rows)

    columns = []
    for i, h in enumerate(headers):
        values = [row[i].strip() if i < len(row) else "" for row in data_rows]
        non_empty = [v for v in values if v]
        col_type = _infer_type(values)
        unique_count = len(set(non_empty))

        col_info = {
            "index": i,
            "name": h,
            "type": col_type,
            "non_null": len(non_empty),
            "unique": unique_count,
            "sample": non_empty[:5],
        }
        columns.append(col_info)

    return {
        "file_rows": n_rows,
        "file_columns": len(headers),
        "columns": columns,
        "sample_rows": [dict(zip(headers, row)) for row in data_rows[:3]],
    }


# ── CSV 출력 ────────────────────────────────────────────────

def write_csv(rows: list[list[str]], output: Path):
    output.parent.mkdir(parents=True, exist_ok=True)
    with open(output, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


# ── CLI ─────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="범용 테이블 데이터 파서 (CSV/xlsx)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python3 parse_tabular.py --file data.csv --summary
  python3 parse_tabular.py --file report.xlsx --summary
  python3 parse_tabular.py --file report.xlsx --list-sheets
  python3 parse_tabular.py --file report.xlsx --sheet "매출" --to-csv output.csv
        """,
    )
    parser.add_argument("--file", required=True, help="입력 파일 경로 (CSV 또는 xlsx)")
    parser.add_argument("--sheet", help="xlsx 시트 이름 (기본: 첫 번째 시트)")
    parser.add_argument("--summary", action="store_true", help="구조 요약 JSON 출력")
    parser.add_argument("--to-csv", help="CSV로 변환 출력할 경로")
    parser.add_argument("--list-sheets", action="store_true", help="xlsx 시트 목록 출력")

    args = parser.parse_args()
    file_path = Path(args.file)

    if not file_path.exists():
        print(f"오류: 파일을 찾을 수 없습니다 — {file_path}", file=sys.stderr)
        sys.exit(1)

    if args.list_sheets:
        sheets = list_sheets(file_path)
        if not sheets:
            print("CSV 파일은 시트 개념이 없습니다.")
        else:
            print("시트 목록:")
            for i, name in enumerate(sheets):
                print(f"  [{i}] {name}")
        sys.exit(0)

    rows = load_tabular(file_path, args.sheet)

    if args.summary:
        summary = summarize(rows)
        summary["source"] = str(file_path)
        summary["parser"] = "stdlib" if file_path.suffix.lower() == ".csv" else "xlsx"
        print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.to_csv:
        out = Path(args.to_csv)
        write_csv(rows, out)
        print(f"CSV 변환 완료: {out} ({len(rows)-1}행)", file=sys.stderr)

    if not args.summary and not args.to_csv:
        # 기본: 요약 출력
        summary = summarize(rows)
        print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
