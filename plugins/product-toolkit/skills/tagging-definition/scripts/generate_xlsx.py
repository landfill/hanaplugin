#!/usr/bin/env python3
"""
태깅 정의서 마크다운 데이터를 읽어 .xlsx 생성.
사용법: python3 generate_xlsx.py <데이터.md> <출력.xlsx>
"""
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent

# 헤더 배경색 (PPT와 동일한 연회색)
HEADER_FILL_COLOR = "E7E6E6"


def _auto_adjust_column_widths(ws):
    """시트의 각 열 너비를 내용에 맞게 자동 조정한다."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is not None:
                # 한글은 약 2배 너비 차지
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


def main():
    if len(sys.argv) < 3:
        print("사용법: python3 generate_xlsx.py <데이터.md> <출력.xlsx>", file=sys.stderr)
        sys.exit(1)
    data_md = Path(sys.argv[1])
    output_xlsx = Path(sys.argv[2])

    sys.path.insert(0, str(SCRIPT_DIR))
    from parse_data_md import parse_md

    data = parse_md(data_md)
    meta = data["meta"]
    screens = data["태깅 대상 화면"]
    events = data["이벤트"]

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("openpyxl이 필요합니다: pip install openpyxl", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.Workbook()
    bold = Font(bold=True)
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 표지
    ws0 = wb.active
    ws0.title = "표지"
    ws0["A1"] = f"{meta['서비스명']} 태깅 정의서"
    ws0["A1"].font = Font(bold=True, size=16)
    ws0["A2"] = meta["작성일"]
    ws0["A3"] = meta["팀"]
    ws0["A4"] = meta["작성자"]
    _auto_adjust_column_widths(ws0)

    # 이력
    ws1 = wb.create_sheet("이력")
    history_labels = ["버전", "변경일", "변경내용", "작성자"]
    history_values = ["0.0.1", meta["작성일"], f"{meta['서비스명']} 태깅 정의 초안", meta["작성자"]]
    for i, (label, val) in enumerate(zip(history_labels, history_values), start=1):
        label_cell = ws1.cell(row=i, column=1, value=label)
        label_cell.font = bold
        label_cell.border = thin_border
        label_cell.fill = header_fill
        val_cell = ws1.cell(row=i, column=2, value=val)
        val_cell.border = thin_border
    _auto_adjust_column_widths(ws1)

    # 개요
    ws2 = wb.create_sheet("개요")
    ws2["A1"] = "태깅 대상 화면"
    ws2["A1"].font = bold
    for i, s in enumerate(screens, start=2):
        ws2[f"A{i}"] = s
    _auto_adjust_column_widths(ws2)

    # 이벤트별 태깅 상세
    for idx, ev in enumerate(events):
        title = ev.get("Event Name (한글)") or ev.get("Event Name (영문)") or f"이벤트{idx+1}"
        event_name_en = ev.get("Event Name (영문)") or ""
        sheet_name = title[:31]  # 엑셀 시트명 길이 제한
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = f"{title}  ({ev.get('호출 시점', '')})"
        ws["A1"].font = Font(bold=True, size=12)
        if event_name_en:
            ws["A2"] = f"Event Name: {event_name_en}"
            ws["A2"].font = Font(italic=True, color="666666")
        headers = ["Event property", "property 구분", "value", "description"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        for r, p in enumerate(ev["properties"], start=4):
            vals = [p["Event property"], p["property 구분"], p["value"], p["description"]]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = thin_border
                if c == 2:  # property 구분은 가운데 정렬
                    cell.alignment = center_align
        _auto_adjust_column_widths(ws)

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_xlsx))
    print(f"저장: {output_xlsx}")


if __name__ == "__main__":
    main()
