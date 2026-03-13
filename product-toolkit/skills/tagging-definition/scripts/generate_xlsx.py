#!/usr/bin/env python3
"""
태깅 정의서 마크다운 데이터를 읽어 .xlsx 생성.
사용법: python3 generate_xlsx.py <데이터.md> <출력.xlsx>
"""
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent


def main():
    if len(sys.argv) < 3:
        print("사용법: python3 generate_xlsx.py <데이터.md> <출력.xlsx>", file=sys.stderr)
        sys.exit(1)
    data_md = Path(sys.argv[1])
    output_xlsx = Path(sys.argv[2])

    sys.path.insert(0, str(SCRIPT_DIR))
    from parse_data_md import parse_md

    data = parse_md(data_md)
    meta = data["meta"]
    screens = data["태깅 대상 화면"]
    events = data["이벤트"]

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        print("openpyxl이 필요합니다: pip install openpyxl", file=sys.stderr)
        sys.exit(2)

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    # 표지
    ws0 = wb.active
    ws0.title = "표지"
    ws0["A1"] = f"{meta['서비스명']} 태깅 정의서"
    ws0["A1"].font = bold
    ws0["A2"] = meta["작성일"]
    ws0["A3"] = meta["팀"]
    ws0["A4"] = meta["작성자"]

    # 이력
    ws1 = wb.create_sheet("이력")
    ws1["A1"] = "버전"
    ws1["B1"] = "0.0.1"
    ws1["A2"] = "변경일"
    ws1["B2"] = meta["작성일"]
    ws1["A3"] = "변경내용"
    ws1["B3"] = f"{meta['서비스명']} 태깅 정의 초안"
    ws1["A4"] = "작성자"
    ws1["B4"] = meta["작성자"]

    # 개요
    ws2 = wb.create_sheet("개요")
    ws2["A1"] = "태깅 대상 화면"
    ws2["A1"].font = bold
    for i, s in enumerate(screens, start=2):
        ws2[f"A{i}"] = s

    # 이벤트별 태깅 상세
    for idx, ev in enumerate(events):
        title = ev.get("Event Name (한글)") or ev.get("Event Name (영문)") or f"이벤트{idx+1}"
        sheet_name = title[:31]  # 엑셀 시트명 길이 제한
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = f"{title}  ({ev.get('호출 시점', '')})"
        ws["A1"].font = bold
        headers = ["Event property", "property 구분", "value", "description"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = bold
        for r, p in enumerate(ev["properties"], start=4):
            ws.cell(row=r, column=1, value=p["Event property"])
            ws.cell(row=r, column=2, value=p["property 구분"])
            ws.cell(row=r, column=3, value=p["value"])
            ws.cell(row=r, column=4, value=p["description"])

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_xlsx))
    print(f"저장: {output_xlsx}")


if __name__ == "__main__":
    main()
